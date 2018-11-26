# Versions:
#   Version 1: Google Spending Data & TotalSC by Day & Report ROI by Day
#   Version 2: From 2 excel files (or csv)
#   Version 3: Export to Google Sheets * Report in Datastudio
#   Version 4: Get Google Spending and TotalSC data from API
#   Version 5: Make this a batch job
#               and make it run everyday, feeding data to Datastudio
#   Next versions: Add Criteo, Facebook, etc, etc.

import os
import Helper
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook

# Google Sheets Integration
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name('ROIReport-101da91f6b59.json', scope)
gc = gspread.authorize(credentials)


def get_trafic_data():
    googleTrafficDataFileName = 'GAReport.xlsx'
    directory = os.path.dirname(__file__) + "/docs/"
    path = directory + googleTrafficDataFileName
    wb = load_workbook(path)

    sheet = wb['Dataset1']
    row_count = sheet.max_row
    column_count = sheet.max_column

    traficData = {}

    i=2
    while i <= row_count:
        

    return {}


# Main Begin

googleSpending = get_google_spending()
googleSC = getGoogleSC()
googleRois = {}
allDates = []

allDates = Helper.get_all_dates()

googleRois = calculateRoi(allDates, googleSpending, googleSC)

wks = gc.open('ROI Report').sheet1
for i in googleRois:
    date = i
    profit = googleSC[i]
    cost = googleSpending[i]
    roi = float(profit / cost)
    wks.append_row([date, profit, cost, roi])

print(wks.get_all_records())

wks.acell('A2').value = '2018-11-04'
print(wks.acell('A2').value)

wks.update_acell('A2', '2018-11-01')
print(wks.get_all_records())
